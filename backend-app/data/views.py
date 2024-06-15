import openpyxl
import datetime

from django.shortcuts import render, HttpResponse

from .models import *

def get_current_smena_childs_list():
    childs = Child.objects.all()
    curent_childs = []

    for child in childs:
        for party in child.party.all():
            if(party.smena.is_curent):
                curent_childs.append(child)
                break

    return curent_childs

def get_current_smena_childs_list_for_party(party_id):
    childs = Child.objects.all()
    curent_childs = []

    for child in childs:
        for party in child.party.all().filter(id=party_id):
            if(party.smena.is_curent):
                curent_childs.append(child)
                break

    return curent_childs

def get_current_smena():
    return Smena.objects.all().filter(is_curent=True).first()

def get_current_smena_party_list():
    return get_current_smena().partys.all()

def get_age(date_of_birth):
    now = datetime.date.today()

    # получаем число полных лет
    years = now.year - date_of_birth.year

    # если день рождения еще не был в этом году, вычтем 1 год
    if date_of_birth.month > now.month:
        # если месяц дня рождения будет позже, то не имеет
        # смысл проверять дату
        years -= 1
    elif date_of_birth.month == now.month:
        if date_of_birth.day > now.day:
            years -= 1

    return years

# Create your views here.
def generate_child_report_handler(request):
    new_file_dest = 'docs/generated/vospitaniki/'
    new_file_name = f'Otchet Vospitaniki {datetime.datetime.now().strftime("%d_%m_%Y %H_%M")}.xlsx'

    obj_workbook = openpyxl.load_workbook('docs/templates/template_02.xlsx')
    obj_sheet = obj_workbook.active


    partys = get_current_smena_party_list()

    for party in partys:
        offset = 0
        obj_sheet_new = obj_workbook.copy_worksheet(obj_sheet)
        obj_sheet_new.title = f'Отряд {party.title}'
        for child in get_current_smena_childs_list_for_party(party.id):
            obj_sheet_new.cell(column=1, row=11+offset).value = offset + 1
            obj_sheet_new.cell(column=2, row=11+offset).value = f'{child.order_state}, {child.order_number}'
            obj_sheet_new.cell(column=3, row=11+offset).value = f'{child.second_name} {child.first_name} {child.third_name}'
            if child.birth_date != None : obj_sheet_new.cell(column=4, row=11+offset).value = f'{child.birth_date.strftime("%d.%m.%Y")}, {child.study_place_full}' 
            if child.birth_date != None : obj_sheet_new.cell(column=5, row=11+offset).value = get_age(child.birth_date)
            obj_sheet_new.cell(column=6, row=11+offset).value = f'{child.live_place_full}, {child.phone_number}'
            obj_sheet_new.cell(column=7, row=11+offset).value = f'{child.mother_full_name}, {child.mother_work} {child.mother_post}'
            obj_sheet_new.cell(column=8, row=11+offset).value = f'{child.father_full_name}, {child.father_work} {child.father_post}'
            if child.social_type != None : obj_sheet_new.cell(column=9, row=11+offset).value = child.social_type.title
            offset+=1

    obj_workbook.save(new_file_dest + new_file_name)

    with open(new_file_dest + new_file_name, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{new_file_name}"'
        return response


def generate_soc_report_handler(request):
    pass

def generate_ticket_report_handler(request):
    new_file_dest = 'docs/generated/putevki/'
    new_file_name = f'Otchet Putevki {datetime.datetime.now().strftime("%d_%m_%Y %H_%M")}.xlsx'

    obj_workbook = openpyxl.load_workbook('docs/templates/template_01.xlsx')
    obj_sheet = obj_workbook.active

    offset = 0
    for child in get_current_smena_childs_list():
        obj_sheet.cell(column=1, row=11+offset).value = child.order_number
        obj_sheet.cell(column=2, row=11+offset).value = child.order_state
        obj_sheet.cell(column=3, row=11+offset).value = f'{child.second_name} {child.first_name} {child.third_name}'
        if child.birth_date != None : obj_sheet.cell(column=4, row=11+offset).value = child.birth_date.strftime("%d.%m.%Y") 
        obj_sheet.cell(column=5, row=11+offset).value = child.study_place_full
        obj_sheet.cell(column=6, row=11+offset).value = get_current_smena().title
        obj_sheet.cell(column=7, row=11+offset).value = f'Мать: {child.mother_full_name}, {child.mother_work} {child.mother_post}; Отец: {child.father_full_name}, {child.father_work} {child.father_post}'
        obj_sheet.cell(column=8, row=11+offset).value = child.live_place_full
        obj_sheet.cell(column=9, row=11+offset).value = f'Мать: {child.mother_phone}; Отец: {child.father_phone}'
        offset+=1

    obj_workbook.save(new_file_dest + new_file_name)

    with open(new_file_dest + new_file_name, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{new_file_name}"'
        return response